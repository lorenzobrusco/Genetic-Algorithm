import random
import math
import sys
import time
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range

if len(sys.argv) < 2:
    _n_cities = 50
else:
    _n_cities = int(sys.argv[1])
_path = "data/dataset-%d.xlsx" % _n_cities
_n_iteration = 10
_travel_cost = 1
_unitary_cost = 1


class Individual:

    def __init__(self, genes=None):
        self._fitness = 0
        if genes is None:
            self._genes = self.generate_individual()
        else:
            self._genes = genes

    def __repr__(self):
        s = ""
        for gene in self._genes:
            s += "|%d" % gene
        s += "|"
        return s

    @staticmethod
    def generate_individual():
        """
            Generate the default individual simple get
            a sorted integers list
        """
        genes = []
        for i in range(_n_cities):
            genes.append(i + 1)
        random.shuffle(genes)
        return genes

    def get_fitness(self):
        """
            Calculate the fitness function
        """
        if self._fitness == 0:
            self._fitness = Fitness.calculate(self._genes)
        return self._fitness

    def get_gene(self, index):
        return self._genes[index]

    def size(self):
        return len(self._genes)

    def set_gene(self, index, value):
        self._genes[index] = value

    def contains(self, value):
        return value in self._genes

    def sort(self):
        return self._genes.sort()


class Population:
    def __init__(self, population_size, initialise=False):
        self._individuals = []
        if initialise is True:
            for i in range(population_size):
                individual = Individual()
                individual.generate_individual()
                self._individuals.append(individual)
        pass

    def get_individual(self, index):
        return self._individuals[index]

    def save_individual(self, individual):
        self._individuals.append(individual)

    def size(self):
        return len(self._individuals)

    def beast_fitness(self):
        """
            It finds the beast individual's fitness
            :return:
        """
        fittest = self.get_individual(0)
        for i in range(self.size()):
            individual = self.get_individual(i)
            if fittest.get_fitness() > individual.get_fitness():
                fittest = individual
        return fittest


class GA:

    def __init__(self, uniform_rate=0.5, mutation_rate=0.015, tournament_size=5, elitism=True):
        self.uniform_rate = uniform_rate
        self.mutation_rate = mutation_rate
        self.tournament_size = tournament_size
        self.elitism = elitism

    def evolve_population(self, population):
        global new_indiv
        global elitism_offset
        new_population = Population(population.size(), False)

        if self.elitism is True:
            new_population.save_individual(population.beast_fitness())
        if self.elitism:
            elitism_offset = 1
        else:
            elitism_offset = 0

        for i in range(elitism_offset, population.size()):
            indiv1 = self.tournament_selection(population)
            indiv2 = self.tournament_selection(population)
            new_indiv = self.order_crossover(indiv1, indiv2)
            new_population.save_individual(new_indiv)

        for i in range(elitism_offset, new_population.size()):
            self.mutate(new_population.get_individual(i))

        return new_population

    def crossover(self, indiv1, indiv2):
        new_sol = Individual()
        for i in range(indiv1.size()):
            if random.randrange(0, 1) <= self.uniform_rate:
                new_sol.set_gene(i, indiv1.get_gene(i))
            else:
                new_sol.set_gene(i, indiv2.get_gene(i))
        return new_sol

    def order_crossover(self, indiv1, indiv2):
        sol = [0 for i in range(0, indiv1.size())]
        new_sol = Individual(sol)
        first_split = random.randint(0, indiv1.size() - 1)
        second_split = random.randint(first_split + 1, indiv1.size())
        index = first_split
        j = second_split % indiv2.size()
        for i in range(first_split, second_split):
            new_sol.set_gene(i, indiv1.get_gene(i))
            index = (i + 1) % indiv1.size()
        while new_sol.contains(0):
            if not new_sol.contains(indiv2.get_gene(j)):
                new_sol.set_gene(index, indiv2.get_gene(j))
                index = (index + 1) % indiv2.size()
            j = (j + 1) % indiv2.size()
        return new_sol

    def mutate(self, indiv):
        for i in range(indiv.size()):
            if random.randrange(0, 1) <= self.mutation_rate:
                value1 = random.randint(0, indiv.size() - 1)
                value2 = random.randint(0, indiv.size() - 1)
                swap1 = indiv.get_gene(value1)
                swap2 = indiv.get_gene(value2)
                indiv.set_gene(value1, swap2)
                indiv.set_gene(value2, swap1)
        pass

    def tournament_selection(self, population):
        tournament = Population(self.tournament_size, False)
        for i in range(self.tournament_size):
            random_id = random.randint(0, population.size() - 1)
            tournament.save_individual(population.get_individual(random_id))
        fittest = tournament.beast_fitness()
        return fittest


class Greedy:

    def __init__(self, solution, sorted=False):
        self.solution = solution
        if sorted is True:
            solution.sort()
        self.already_visited = []

    def solve(self):
        last = 0
        self.already_visited.append(self.solution.get_gene(last))
        while len(self.already_visited) != self.solution.size():
            last = self.search_best_local_solution(last)

    def search_best_local_solution(self, row):
        wb = load_workbook(filename=_path)
        profit = 0
        profits = []
        for index in range(self.solution.size()):
            cost = 0
            cost += wb["Distance"].cell(row=self.solution.get_gene(row),
                                        column=self.solution.get_gene(index)).value * _travel_cost
            profits.append((self.solution.get_gene(index), cost, index))
            profits.sort(key=lambda tup: tup[1])
        for i in range(len(profits)):
            if profits[i][0] not in self.already_visited:
                self.already_visited.append(profits[i][0])
                return profits[i][2]

        return 0

    def fitness(self):
        return Fitness.calculate(self.already_visited)

    def beast_fitness(self):
        individual = Individual(self.already_visited)
        return individual


class GRASP:
    def __init__(self, n_cities, iteration):
        self.n_cities = n_cities
        self.iteration = iteration
        self.restricted_candidate_list = []

    def solve(self, printed=False):
        for i in range(self.iteration):
            population = Population(self.n_cities, True)
            local_greedy = Greedy(population.beast_fitness(), sorted=False)
            local_greedy.solve()
            solution = local_greedy.beast_fitness()
            if printed is True:
                print("Iteration %d: generate new solution,  fitness %d" % (
                    i, local_greedy.beast_fitness().get_fitness()))
            self.restricted_candidate_list.append((solution, solution.get_fitness()))
        self.restricted_candidate_list.sort(key=lambda tup: tup[1], reverse=True)

    def beast_fitness(self):
        return self.restricted_candidate_list[0][0]


class Fitness:

    def __init__(self):
        self.solution = []

    def set_solution(self, solution):
        self.solution = solution

    @staticmethod
    def calculate(solution):
        if solution is None:
            return 0
        wb = load_workbook(filename=_path)
        cost = 0
        for index in range(len(solution) - 1):
            cost += wb["Distance"].cell(row=solution[index], column=solution[index + 1]).value * _travel_cost
        return cost


def create_file(path, n_cities):
    """
        :param path:
        :param n_cities:
        :return:
    """
    wb = Workbook()
    ws = wb.create_sheet("Location")
    ws1 = wb.create_sheet("Distance")
    ws2 = wb.create_sheet("Cost")
    ws3 = wb.create_sheet("Profit")
    location = []
    for i in range(1, n_cities + 1):
        x = random.randrange(0, 1000)
        y = random.randrange(0, 1000)
        ws.cell(row=i, column=1, value=x)
        ws.cell(row=i, column=2, value=y)
        location.append((x, y))

    for i in range(0, n_cities):
        value = random.randint(1, 100)
        ws2.cell(row=i + 1, column=1, value=value)
        value = random.randint(100, 1000)
        ws3.cell(row=i + 1, column=1, value=value)
        for j in range(0, n_cities):

            x_distance = abs(location[i][0] - location[j][0])
            y_distance = abs(location[i][1] - location[j][1])
            distance = math.sqrt((x_distance * x_distance) + (y_distance * y_distance))
            ws1.cell(row=i + 1, column=j + 1, value=distance)
            if i == j:
                ws1.cell(row=i + 1, column=j + 1, value=0)
    ws1.cell(row=1, column=1, value=0)
    del wb['Sheet']
    wb.save(path)


def read_data(path, n_cities):
    """
        :param path:
        :param n_cities:
        :return:
    """
    file = Path(path)
    if not file.exists():
        print("Creating file...")
        create_file(path, n_cities)
    location = []
    wb = load_workbook(filename=path)
    ws1 = wb["Location"]

    for row in range(n_cities):
        x = ws1.cell(row=row + 1, column=1).value
        y = ws1.cell(row=row + 1, column=2).value
        location.append((x, y))

    return location


def generate_graph(graph, location, show=True):
    G = nx.DiGraph()
    for i in range(len(location)):
        x = location[i][0]
        y = location[i][1]
        G.add_node(graph.get_gene(i), pos=(x, y))
    for i in range(graph.size() - 1):
        G.add_edge(graph.get_gene(i), graph.get_gene(i + 1))
    pos = nx.get_node_attributes(G, 'pos')
    nx.draw_networkx_nodes(G, pos, cmap=plt.get_cmap('jet'), node_size=1000, alpha=0.7)
    nx.draw_networkx_labels(G, pos, font_color='white')
    nx.draw_networkx_edges(G, pos, arrows=True)
    if show is True:
        plt.show()


def solve_greedy():
    population = Population(_n_cities, True)
    greedy = Greedy(population.beast_fitness(), True)
    start = time.clock()
    greedy.solve()
    end = time.clock()
    greedy_time = (end - start)
    print("Best solution with GREEDY: %d time %d\n" % (greedy.fitness(), greedy_time), greedy.beast_fitness())


def solve_ga(graphic=False, printed=False):
    location = read_data(_path, _n_cities)
    population = Population(_n_cities, True)
    ga = GA()
    start = time.clock()
    for i in range(_n_iteration):
        if printed is True:
            print("Iteration %d: generate new population,  fitness %d" % (i, population.beast_fitness().get_fitness()))
        population = ga.evolve_population(population)
    end = time.clock()
    ga_time = (end - start)
    print("Best solution with GA: %d time: %d\n" % (population.beast_fitness().get_fitness(), ga_time),
          population.beast_fitness())
    if graphic is True:
        generate_graph(population.beast_fitness(), location)


def solve_grasp(printed=False):
    start = time.clock()
    grasp = GRASP(_n_cities, _n_iteration)
    grasp.solve(printed=printed)
    end = time.clock()
    grasp_time = (end - start)
    print("Best solution with GRASP: %d time: %d\n" % (grasp.beast_fitness().get_fitness(), grasp_time),
          grasp.beast_fitness())


if __name__ == "__main__":
    cities = [index for index in range(1, _n_cities + 1)]
    fitness = Fitness.calculate(cities)
    print("Initial fitness is %d" % fitness)
    solve_greedy()
    solve_ga(printed=True)
    solve_grasp()
