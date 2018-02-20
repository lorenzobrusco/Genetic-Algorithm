import random
import math
import sys
import copy
import time
import numpy as np
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.compat import range

if len(sys.argv) < 2:
    _n_cities = 51
else:
    _n_cities = int(sys.argv[1])
_path = "data/dataset-HP.xlsx"
_n_iteration = 100
_travel_cost = 1
_unitary_cost = 1
_unused_node = -9


def calculate_distance(location):
    distances = []
    for i in range(len(location)):
        for j in range(len(location)):
            x_distance = abs(location[i][0] - location[j][0])
            y_distance = abs(location[i][1] - location[j][1])
            distance = math.sqrt((x_distance * x_distance) + (y_distance * y_distance))
            distances.append(distance)
    distances = np.reshape(distances, (len(location), len(location)))
    return distances


def read_data(path, n_cities):
    """
        :param path:
        :param n_cities:
        :return:
    """
    file = Path(path)
    if not file.exists():
        print("No file found")
        exit(-1)
    nodes = []
    wb = load_workbook(filename=path)
    ws1 = wb["eil%d" % n_cities]

    for row in range(0, n_cities):
        x = ws1.cell(row=row + 1, column=2).value
        y = ws1.cell(row=row + 1, column=3).value
        prof = ws1.cell(row=row + 1, column=4).value
        nodes.append((x, y, prof))

    return nodes


_nodes = read_data(_path, _n_cities)
_distance = calculate_distance(_nodes)


class Individual:

    def __init__(self, genes=None):
        self._fitness = None
        if genes is None:
            self._genes = self.generate_individual()
        else:
            self._genes = genes
        self.calculate_fitness()

    def __repr__(self):
        s = ""
        for i in range(self.size()):
            s += "|%2d" % i
        s += "|\n "
        for gene in self._genes:
            s += "|%2d" % gene
        s += "|"
        return s

    @staticmethod
    def generate_individual():
        """
            Generate the default individual simple get
            a sorted integers list
        """
        genes = []
        for i in range(_n_cities):
            genes.append((i + 1) % _n_cities)
        return genes

    def calculate_total_distance(self):
        index = 0
        distance = 0
        stop = False
        for i in range(len(self._genes)):
            distance += _distance[index][self._genes[index]]
            index = self._genes[index]
            if stop:
                break
            if self._genes == 0:
                stop = True
        return distance

    def calculate_fitness(self):
        """
            Calculate the fitness function
        """
        self._fitness = Fitness.calculate(self._genes)

    def get_fitness(self):
        return self._fitness

    def get_genes(self):
        return self._genes

    def get_gene(self, index):
        return self._genes[index]

    def size(self):
        return len(self._genes)

    def set_gene(self, index, value):
        self._genes[index] = value

    def contains(self, value):
        return value in self._genes

    def sort(self):
        return self._genes.sort()


class Population:
    def __init__(self, population_size, initialise=False):
        self._individuals = []
        if initialise is True:
            for i in range(population_size):
                individual = Individual()
                individual.generate_individual()
                self._individuals.append(individual)
        pass

    def get_individual(self, index):
        return self._individuals[index]

    def save_individual(self, individual):
        self._individuals.append(individual)

    def size(self):
        return len(self._individuals)

    def beast_fitness(self):
        """
            It finds the beast individual's fitness
            :return:
        """
        fittest = self.get_individual(0)
        for i in range(self.size()):
            individual = self.get_individual(i)
            if fittest.get_fitness() <= individual.get_fitness():
                fittest = individual
        return fittest


class GA:

    def __init__(self, uniform_rate=0.5, mutation_rate=0.015, tournament_size=10, elitism=True, twopt=False):
        self.twopt = twopt
        self.uniform_rate = uniform_rate
        self.mutation_rate = mutation_rate
        self.tournament_size = tournament_size
        self.elitism = elitism
        self.best = Individual()

    def evolve_population(self, population):
        new_population = Population(population.size(), initialise=False)
        if self.elitism:
            if self.best.get_fitness() <= population.beast_fitness().get_fitness():
                if self.twopt is True:
                    self.best = self.two_opt(population.beast_fitness())
                else:
                    self.best = population.beast_fitness()
        for i in range(population.size()):
            if self.elitism:
                indiv1 = copy.deepcopy(self.best)
            else:
                indiv1 = self.tournament_selection(population)
            indiv2 = self.tournament_selection(population)
            new_indiv = self.crossover(indiv1, indiv2)
            new_population.save_individual(new_indiv)

        for i in range(new_population.size()):
            self.mutate(new_population.get_individual(i))
        return new_population

    def crossover(self, indiv1, indiv2):
        labeled = []
        index = 0
        if indiv1.get_fitness() >= indiv2.get_fitness():
            new_sol = indiv1
            for i in range(new_sol.size()):
                if not new_sol.contains(indiv2.get_gene(i)) and indiv2.get_gene(i) != _unused_node:
                    labeled.append(indiv2.get_gene(i))
        else:
            new_sol = indiv2
            for i in range(new_sol.size()):
                if not new_sol.contains(indiv1.get_gene(i)) and indiv1.get_gene(i) != _unused_node:
                    labeled.append(indiv1.get_gene(i))

        for i in range(new_sol.size()):
            if new_sol.get_gene(i) == 0:
                index = i
        for i in range(len(labeled)):
            new_sol.set_gene(index, labeled[i])
            new_sol.set_gene(labeled[i], 0)
            index = labeled[i]
        return new_sol

    def mutate(self, indiv):
        for i in range(indiv.size()):
            if random.uniform(0, 1) <= self.mutation_rate:
                if indiv.get_gene(i) == _unused_node:
                    self.__add_nodes(i, indiv)
                else:
                    self.__remove_nodes(i, indiv)
        indiv.calculate_fitness()

    def tournament_selection(self, population):
        tournament = Population(self.tournament_size, False)
        for k in range(self.tournament_size):
            random_id = random.randint(0, population.size() - 1)
            tournament.save_individual(population.get_individual(random_id))
        fittest = tournament.beast_fitness()
        return fittest

    def __add_nodes(self, i, indiv):
        index = 0
        for j in range(indiv.size()):
            index = random.randint(1, indiv.size() - 1)
            if indiv.get_gene(index) != _unused_node:
                break
            else:
                index = _unused_node
        if index != _unused_node:
            swap_value_new = indiv.get_gene(index)
            indiv.set_gene(i, swap_value_new)
            indiv.set_gene(index, i)

    def __remove_nodes(self, i, indiv):
        swap_value = indiv.get_gene(i)
        new_index = -1
        for j in range(indiv.size()):
            if i == indiv.get_gene(j):
                new_index = j
                break
        if new_index != i and new_index != -1 and swap_value != new_index and i != 0:
            indiv.set_gene(new_index, swap_value)
            indiv.set_gene(i, _unused_node)

    def two_opt(self, solution):
        _break = False
        size = solution.size()
        new_tour = copy.deepcopy(solution)
        beast_distance = solution.calculate_total_distance()
        while _break == False:
            for i in range(1, size - 1):
                for j in range(i + 1, size):
                    if new_tour.get_gene(i) == _unused_node or new_tour.get_gene(j) == _unused_node:
                        continue
                    new_tour = self.two_opt_swap(solution, new_tour, i, j)
                    local_distance = new_tour.calculate_total_distance()
                    if local_distance < beast_distance:
                        solution = copy.deepcopy(new_tour)
                        beast_distance = local_distance
                        _break = True

        return solution

    def two_opt_swap(self, solution, tour, i, j):
        index = i
        local_tour_index = []
        local_tour = []
        print("[BEFORE]", tour)
        for k in range(solution.size()):
            local_tour_index.append(index)
            local_tour.append(solution.get_gene(index))
            index = solution.get_gene(index)
            j -= 1
            if j <= 0:
                break
        local_tour.reverse()
        for k in range(len(local_tour)):
            tour.set_gene(local_tour_index[k], local_tour[k])
        print("[AFTER]", tour)
        return tour


class Fitness:

    @staticmethod
    def calculate(solution, printed=False):
        if solution is None:
            return 0
        prof = 0
        index = 0
        stop = False
        label = []
        for i in range(len(solution)):
            if solution[index] == _unused_node:
                return -9999
            cost = _distance[index][solution[index]] * _travel_cost
            prof += (_nodes[solution[index]][2] - cost)
            index = solution[index]
            label.append(solution[index])
            if printed:
                print(" FROM %d ---> %d COST %d PROF %d" % (index, solution[index], cost, prof))
            if stop:
                break
            if solution[index] == 0:
                stop = True

        for i in range(len(solution)):
            if solution[i] not in label:
                #prof -= _nodes[solution[i]][2]
                solution[i] = _unused_node

        return prof


def generate_graph(graph, location, show=True):
    plt.figure(figsize=(9, 9))
    G = nx.DiGraph()
    for i in range(len(location)):
        x = location[i][0]
        y = location[i][1]
        G.add_node(i, pos=(x, y))
    index = 0
    stop = False
    for i in range(graph.size()):
        G.add_edge(index, graph.get_gene(index))
        index = graph.get_gene(index)
        if stop:
            break
        if graph.get_gene(index) == 0:
            stop = True
    pos = nx.get_node_attributes(G, 'pos')
    nx.draw_networkx_nodes(G, pos, cmap=plt.get_cmap('jet'))
    nx.draw_networkx_labels(G, pos)
    nx.draw_networkx_edges(G, pos, arrows=True)

    if show is True:
        plt.show()


def solve_ga(graphic=False, printed=False):
    location = read_data(_path, _n_cities)
    population = Population(_n_cities, True)
    #ga = GA(twopt=True)
    ga = GA(twopt=False)
    start = time.clock()

    for i in range(_n_iteration):
        if printed is True:
            print(
                "Iteration %d: generate new population,  fitness %d\n" % (i, population.beast_fitness().get_fitness()))
        population = ga.evolve_population(population)
    end = time.clock()
    ga_time = (end - start)
    print("Best solution with GA: %d time: %d\n" % (population.beast_fitness().get_fitness(), ga_time),
          population.beast_fitness())
    if graphic is True:
        generate_graph(population.beast_fitness(), location)


if __name__ == "__main__":
    cities = [(index + 1) % _n_cities for index in range(0, _n_cities)]
    fitness = Fitness.calculate(cities)
    print("Initial fitness is %d" % fitness)
    solve_ga(graphic=True, printed=True)
