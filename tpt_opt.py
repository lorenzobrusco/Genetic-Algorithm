import random
import math
import sys
import copy
import time
import numpy as np
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.compat import range
from threading import Thread, Event
import time

if len(sys.argv) < 2:
    _n_cities = 51
else:
    _n_cities = int(sys.argv[1])
_path = "data/dataset-HP.xlsx"
_n_iteration = 100
_depo = 0
_travel_cost = 1
_unitary_cost = 1
_unused_node = -9
_penality = False
_maximum_dinstance_from_depot = 0
_minimum_dinstance_from_depot = 0
random.seed(7)
stop_event = Event()

def calculate_distance(location):
    distances = []
    global _maximum_dinstance_from_depot
    global _minimum_dinstance_from_depot_
    assign = True

    for i in range(len(location)):
        for j in range(len(location)):
            x_distance = abs(location[i][0] - location[j][0])
            y_distance = abs(location[i][1] - location[j][1])
            distance = math.sqrt((x_distance * x_distance) + (y_distance * y_distance))
            if(distance > 0):

                if assign:
                    _minimum_dinstance_from_depot =\
                    _maximum_dinstance_from_depot = distance
                    assign = False
                else:

                    if(distance < _minimum_dinstance_from_depot):

                        _minimum_dinstance_from_depot = distance

                    if(distance > _maximum_dinstance_from_depot):
                        _maximum_dinstance_from_depot = distance

            distances.append(distance)
    distances = np.reshape(distances, (len(location), len(location)))
    return distances

'''
            This function is the implementation of :
"Heuristics for the traveling repairman problem with profits"
T. Dewilde, D. Cattrysse , S. Coene , F.C.R. Spieksma, P. Vansteenwegen
                            -Section 4

'''

def calculate_profits(location):
    #random_integers [low,high]  we need (low,high] low excluded

    for i in range(_n_cities):
        profit =\
        np.random.random_integers((_minimum_dinstance_from_depot +1 ),(_n_cities/2)*_maximum_dinstance_from_depot)
        location[i][2] = profit
        #print ("Node %d  profit %f " % (i,profit))

def read_data(path, n_cities):
    """
        :param path:
        :param n_cities:
        :return:
    """
    file = Path(path)
    if not file.exists():
        print("No file found")
        exit(-1)
    nodes = []
    wb = load_workbook(filename=path)
    ws1 = wb["eil%d" % n_cities]

    for row in range(n_cities):
        x = ws1.cell(row=row + 1, column=2).value
        y = ws1.cell(row=row + 1, column=3).value
        nodes.append([x, y, 0]) # [] see it as a list () as a tuple

    return nodes

_nodes = read_data(_path, _n_cities)
_distance = calculate_distance(_nodes)
calculate_profits(_nodes)

class Individual:

    def __init__(self, create=False, genes=None):
        self._fitness = None
        self._genes = []
        self.useless = []
        if create is True:
            self._genes = self.generate_individual()
        if genes is not None:
            self._genes = genes
        self.calculate_fitness()

    def __repr__(self):
        s = "["
        for gene in self._genes:
            s += " %2d" % gene
        s += "]"
        return s

    def generate_individual(self):
        """
            Generate the default individual simple get
            a sorted integers list
        """
        genes = []
        for i in range(_n_cities):
            genes.append((i + 1) % _n_cities)
        return genes

    def calculate_total_distance(self):
        distance = 0
        for i in range(len(self._genes)):
            distance += _distance[self._genes[i]][self._genes[(i + 1) % len(self._genes)]]
        return distance

    def calculate_fitness(self):
        """
            Calculate the fitness function
        """
        self._fitness = Fitness.calculate(self._genes)

    def remove(self, i):
        useless = self._genes[i]
        self._genes.remove(useless)
        self.useless.append(useless)

    def get_fitness(self):
        return self._fitness

    def get_genes(self):
        return self._genes

    def set_genes(self, genes):
        self._genes = genes

    def get_gene(self, index):
        return self._genes[index]

    def get_useless(self):
        return self.useless

    def size(self):
        return len(self._genes)

    def append(self, value):
        self._genes.append(value)

    def set_gene(self, index, value):
        self._genes[index] = value

    def contains(self, value):
        return value in self._genes

    def sort(self):
        return self._genes.sort()

    def insert(self, i, value):
        self._genes.insert(i, value)


class Population:
    def __init__(self, population_size, initialise=False):
        self._individuals = []
        if initialise is True:
            for i in range(population_size):
                individual = Individual(create=True)
                self._individuals.append(individual)
        pass

    def get_individual(self, index):
        return self._individuals[index]

    def save_individual(self, individual):
        self._individuals.append(individual)

    def size(self):
        return len(self._individuals)

    def beast_fitness(self):
        """
            It finds the beast individual's fitness
            :return:
        """
        fittest = self.get_individual(0)
        for i in range(self.size()):
            individual = self.get_individual(i)
            if fittest.get_fitness() <= individual.get_fitness():
                fittest = individual
        return fittest


class GA:

    def __init__(self, uniform_rate=0.5, mutation_rate=0.015, tournament_size=10, elitism=True, two_opt=False):
        self.uniform_rate = uniform_rate
        self.mutation_rate = mutation_rate
        self.tournament_size = tournament_size
        self.elitism = elitism
        self.two_opt = two_opt
        self.best = Individual(True)

    def evolve_population(self, population):
        new_population = Population(population.size(), initialise=False)
        if self.elitism:
            if self.best.get_fitness() <= population.beast_fitness().get_fitness():
                if self.two_opt is True:
                    self.best.set_genes(self.twoOpt(population.beast_fitness().get_genes()))
                else:
                    self.best = population.beast_fitness()
        for i in range(population.size()):
            if self.elitism:
                indiv1 = copy.deepcopy(self.best)
            else:
                indiv1 = self.tournament_selection(population)
            indiv2 = self.tournament_selection(population)
            new_indiv = self.crossover(indiv1, indiv2)
            new_population.save_individual(new_indiv)

        for i in range(new_population.size()):
            self.mutate(new_population.get_individual(i))
        return new_population

    def crossover(self, indiv1, indiv2):
        labeled = []
        new_indiv = Individual()
        new_indiv.insert(0, _depo)
        if random.uniform(0, 1) <= self.uniform_rate:
            pather1 = indiv1
            pather2 = indiv2
        else:
            pather2 = indiv1
            pather1 = indiv2
        first_split = random.randint(0, pather1.size() - 2)
        second_split = random.randint(first_split + 1, pather1.size() - 1)
        for i in range(first_split, second_split):
            labeled.append(pather1.get_gene(i))
        for i in range(pather2.size()):
            if pather2.get_gene(i) not in labeled:
                if i < first_split:
                    new_indiv.append(pather2.get_gene(i))
                else:
                    labeled.append(pather2.get_gene(i))
        for i in range(len(labeled)):
            new_indiv.append(labeled[i])
            first_split += 1
        return new_indiv

    def mutate(self, indiv):
        for i in range(1, indiv.size()):
            if i >= indiv.size():
                break
            if random.uniform(0, 1) <= self.mutation_rate:
                if random.uniform(0, 1) <= self.uniform_rate:
                    indiv.remove(i)
                else:
                    if len(indiv.get_useless()) > 0:
                        useless = indiv.get_useless().pop()
                        indiv.insert(i, useless)
                    else:
                        indiv.remove(i)
        uniq = []
        [uniq.append(x) for x in indiv.get_genes() if x not in uniq]
        indiv.set_genes(uniq)
        indiv.append(_depo)
        indiv.calculate_fitness()

    def tournament_selection(self, population):
        tournament = Population(self.tournament_size, False)
        for k in range(self.tournament_size):
            random_id = random.randint(0, population.size() - 1)
            tournament.save_individual(population.get_individual(random_id))
        fittest = tournament.beast_fitness()
        return fittest

    def twoOpt(self, route):
        xx = 0
        while (True):
            xx += 1
            temp_route = list(route)
            route_distance = -999999999
            for i in range(1, len(route) - 2):
                for j in range(i + 1, len(route) - 1):
                    new_route = route[:i] + list(reversed(route[i:j + 1])) + route[j + 1:]
                    diff_distance = _distance[route[i - 1]][route[i]] + _distance[route[j]][route[j + 1]]
                    diff_distance = diff_distance - _distance[new_route[i - 1]][new_route[i]] - _distance[new_route[j]][
                        new_route[j + 1]]
                    if diff_distance > route_distance:
                        temp_route = list(new_route)
                        route_distance = diff_distance
            if route_distance > 0.01:
                route = list(temp_route)
            else:
                break
        return route


class Fitness:

    @staticmethod
    def calculate(solution):
        if solution is None or len(solution) == 0:
            return -999999
        prof = 0
        for i in range(len(solution)):
            cost = _distance[solution[i]][solution[(i + 1) % len(solution)]] * _travel_cost
            prof += (_nodes[solution[(i + 1) % len(solution)]][2] - cost)
        if _penality is True:
            nodes = []
            [nodes.append(x) for x in range(_n_cities) if x not in solution]
            for i in nodes:
                prof -= _nodes[i][2]
        return prof


def generate_graph(graph, location, show=True):
    G = nx.DiGraph()
    plt.figure(figsize=(9, 9))
    plt.axis('off')
    for i in range(len(location)):
        x = location[i][0]
        y = location[i][1]
        G.add_node(i, pos=(x, y))
    for i in range(graph.size()):
        G.add_edge(graph.get_gene(i), graph.get_gene((i + 1) % graph.size()))
    pos = nx.get_node_attributes(G, 'pos')
    nx.draw_networkx_nodes(G, pos, cmap=plt.get_cmap('jet'))
    nx.draw_networkx_labels(G, pos)
    nx.draw_networkx_edges(G, pos, arrows=True)

    if show is True:
        plt.show()

def action(printed=False,ga = None,population = None):

    print ("Algorithm started")
    for i in range(_n_iteration):
        if printed is True:
            print("Iteration %d: generate new population,  fitness %d\n" % (i, population.beast_fitness().get_fitness()))
        population = ga.evolve_population(population)
        if stop_event.is_set():
            print ("Timeout reached")
            break
    stop_event.set()

def solve_ga(graphic=False, printed=False):
    location = read_data(_path, _n_cities)
    population = Population(_n_cities, True)
    ga = GA(two_opt=True)

    start = time.clock()

    action_thread = Thread (target=action, args=(printed,ga,population))
    action_thread.start()
    action_thread.join(timeout=int(sys.argv[2]))
    end = time.clock()
    ga_time = (end - start)
    print("\nInstance: ")
    print("eil" + str(_n_cities))
    print()

    print("Best Objective Value:")
    print("%.2f" % population.beast_fitness().get_fitness())
    print()

    print("Number of Customers Visited (Depot Excluded):")
    print(population.beast_fitness().size() - 2)
    print()

    print("Sequence of Customers Visited:")
    print(population.beast_fitness())
    print()

    print("CPU Time (s):")
    print("%.2f" % ga_time)

    if graphic is True:
        generate_graph(population.beast_fitness(), location)


if __name__ == "__main__":
    cities = [(index + 1) % _n_cities for index in range(0, _n_cities)]
    fitness = Fitness.calculate(cities)
    solve_ga(graphic=True, printed=False)
